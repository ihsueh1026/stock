"""
Fetch TWSE daily price history for one or more stocks plus the
TAIEX index, compute common Taiwan-market technical indicators,
and save the most recent N trading days as an Excel (.xlsx) file
with one worksheet per stock.

Output columns (in order):
    Date, TAIEX, High, Low, Close, Volume (lots), Change %,
    MA5, MA10, MA20, RSI(6), RSI(12), KD-K, KD-D, MACD-OSC

Indicator definitions (Taiwan-market conventions):
    MAn        : simple moving average of close, n days
    Change %   : (Close - PrevClose) / PrevClose * 100
    RSI(n)     : Wilder-smoothed RSI on close-to-close changes
    KD (9,3,3) : RSV uses 9-day intraday High/Low (consistent with
                 most pro charting tools, e.g. TradingView).
                 K = 2/3 * prevK + 1/3 * RSV
                 D = 2/3 * prevD + 1/3 * K
                 K0 = D0 = 50 (standard seed value)
    MACD       : EMA12 - EMA26 = DIF
                 signal = EMA9(DIF)
                 OSC = DIF - signal  (often called MACD histogram)

To make every output row's indicators meaningful, the script
fetches a warmup buffer (default 60 trading days) BEFORE the
output window, computes indicators across the whole series, and
then trims the result down to the requested rows.

Usage:
    pip install requests openpyxl
    python fetch_twse_daily.py                       # 2324, 30 days
    python fetch_twse_daily.py -s 2330 2317 -n 60
    python fetch_twse_daily.py --no-taiex            # skip TAIEX column
"""

import argparse
import os
import sys
import time
from datetime import datetime, date
from typing import List, Optional, Callable, Dict, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ---- Constants -------------------------------------------------------------

STOCK_URL = "https://www.twse.com.tw/exchangeReport/STOCK_DAY"
TAIEX_URL = "https://www.twse.com.tw/exchangeReport/FMTQIK"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

REQUEST_INTERVAL_SEC = 3
MAX_LOOKBACK_MONTHS = 24

# Retry settings for transient TWSE errors. Sometimes the API returns
# bogus "stat" messages (e.g. "查詢日期大於今日" for a past month) when
# it's rate-limited or hiccuping. We retry with exponential-ish backoff.
MAX_RETRIES = 3
RETRY_BACKOFF_SEC = [5, 10, 20]   # one entry per retry attempt

# Extra trading days to fetch before the output window so that
# indicators are warmed up by the time we reach the first output row.
# 250 ≈ 1 trading year — long enough for the 26-day EMA inside MACD
# to converge and align closely with what broker apps display.
# (Broker apps load multi-year history, so their EMAs are even more
# settled; the longer our warmup, the smaller the residual difference.)
WARMUP_DAYS = 250

# STOCK_DAY column indices.
STOCK_IDX_DATE = 0
STOCK_IDX_VOLUME = 1
STOCK_IDX_HIGH = 4
STOCK_IDX_LOW = 5
STOCK_IDX_CLOSE = 6

# FMTQIK column indices (TAIEX daily).
TAIEX_IDX_DATE = 0
TAIEX_IDX_CLOSE = 4


# ---- TWSE fetch ------------------------------------------------------------

def _is_real_future_date(label: str) -> bool:
    """
    Decide if a 'date > today' error from TWSE is genuine. We compare
    the YYYY-MM in `label` to the current year/month: only months strictly
    after the current month should ever legitimately get that error.
    Anything else is a transient API hiccup we should retry.
    """
    try:
        y_str, m_str = label.split("-")
        year, month = int(y_str), int(m_str)
    except (ValueError, AttributeError):
        return False
    now = datetime.now()
    return (year, month) > (now.year, now.month)


def _fetch(url: str, params: dict, label: str) -> List[List[str]]:
    """
    Fetch one month of data with retry on transient errors.

    TWSE sometimes returns weird stat messages (e.g. claiming a past
    month is "in the future", or empty data when there should be some)
    under load. We retry these with backoff.
    """
    last_stat = None
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = requests.get(url, params=params, headers=HEADERS, timeout=15)
            resp.raise_for_status()
            payload = resp.json()
        except (requests.RequestException, ValueError) as exc:
            # Network error or non-JSON body — definitely retry.
            last_stat = f"request error: {exc}"
            if attempt < MAX_RETRIES:
                wait = RETRY_BACKOFF_SEC[attempt]
                print(f"\n      [retry {attempt + 1}/{MAX_RETRIES}] {label} "
                      f"({last_stat}); waiting {wait}s...", end="")
                time.sleep(wait)
                continue
            print(f"\n      [fail] {label}: {last_stat}", end=" ")
            return []

        stat = payload.get("stat")
        if stat == "OK":
            return payload.get("data", [])

        last_stat = stat
        # Real "future date" — no point retrying, the data doesn't exist yet.
        if stat and "大於今日" in stat and _is_real_future_date(label):
            print(f"\n    [skip] {label}: {stat}", end=" ")
            return []
        # Otherwise it's likely transient. Retry.
        if attempt < MAX_RETRIES:
            wait = RETRY_BACKOFF_SEC[attempt]
            print(f"\n      [retry {attempt + 1}/{MAX_RETRIES}] {label} "
                  f"({stat}); waiting {wait}s...", end="")
            time.sleep(wait)
            continue

    print(f"\n    [skip] {label}: {last_stat} (after {MAX_RETRIES} retries)", end=" ")
    return []


def fetch_stock_month(year: int, month: int, stock_no: str) -> List[List[str]]:
    return _fetch(
        STOCK_URL,
        {"response": "json", "date": f"{year}{month:02d}01", "stockNo": stock_no},
        f"{year}-{month:02d}",
    )


def fetch_taiex_month(year: int, month: int) -> List[List[str]]:
    return _fetch(
        TAIEX_URL,
        {"response": "json", "date": f"{year}{month:02d}01"},
        f"{year}-{month:02d}",
    )


def previous_month(year: int, month: int) -> Tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


def fetch_recent_rows(
    fetch_one_month: Callable[[int, int], List[List[str]]],
    target_rows: int,
    label: str,
) -> List[List[str]]:
    """Walk backwards month-by-month until at least target_rows are collected."""
    collected: List[List[str]] = []
    now = datetime.now()
    year, month = now.year, now.month

    for _ in range(MAX_LOOKBACK_MONTHS):
        print(f"    -> {year}-{month:02d}", end=" ")
        rows = fetch_one_month(year, month)
        print(f"({len(rows)} rows)")
        collected = rows + collected   # API gives oldest-first within a month
        time.sleep(REQUEST_INTERVAL_SEC)

        if len(collected) >= target_rows:
            break
        year, month = previous_month(year, month)
    else:
        print(f"    [warn] hit MAX_LOOKBACK_MONTHS for {label}; "
              f"only got {len(collected)} rows")

    return collected[-target_rows:] if len(collected) > target_rows else collected


# ---- Conversion helpers ----------------------------------------------------

def roc_to_western_date(roc_str: str) -> Optional[date]:
    try:
        roc_y, m, d = roc_str.strip().split("/")
        return date(int(roc_y) + 1911, int(m), int(d))
    except (ValueError, AttributeError):
        return None


def to_number(raw: str, *, integer: bool = False):
    cleaned = raw.replace(",", "").strip()
    if cleaned in ("", "--", "X"):
        return None
    try:
        return int(cleaned) if integer else float(cleaned)
    except ValueError:
        return None


# ---- Technical indicators --------------------------------------------------

def sma(values: List[Optional[float]], n: int) -> List[Optional[float]]:
    """Simple moving average. Returns None until n full values are available."""
    out: List[Optional[float]] = []
    window: List[float] = []
    for v in values:
        if v is None:
            window = []          # reset on gap; conservative
            out.append(None)
            continue
        window.append(v)
        if len(window) > n:
            window.pop(0)
        out.append(sum(window) / n if len(window) == n else None)
    return out


def rsi_wilder(closes: List[Optional[float]], n: int) -> List[Optional[float]]:
    """
    Wilder-smoothed RSI on close-to-close changes.
    First value at index n (uses simple average of first n gains/losses),
    then exponentially smoothed: avg = (avg*(n-1) + current) / n.
    """
    out: List[Optional[float]] = [None] * len(closes)
    if len(closes) <= n:
        return out

    gains: List[float] = []
    losses: List[float] = []
    for i in range(1, len(closes)):
        prev, curr = closes[i - 1], closes[i]
        if prev is None or curr is None:
            gains.append(0.0)
            losses.append(0.0)
            continue
        diff = curr - prev
        gains.append(max(diff, 0.0))
        losses.append(max(-diff, 0.0))

    # First averages are simple means of the first n diffs.
    avg_gain = sum(gains[:n]) / n
    avg_loss = sum(losses[:n]) / n
    out[n] = _rsi_value(avg_gain, avg_loss)

    # Subsequent values use Wilder smoothing.
    for i in range(n + 1, len(closes)):
        g = gains[i - 1]
        l = losses[i - 1]
        avg_gain = (avg_gain * (n - 1) + g) / n
        avg_loss = (avg_loss * (n - 1) + l) / n
        out[i] = _rsi_value(avg_gain, avg_loss)
    return out


def _rsi_value(avg_gain: float, avg_loss: float) -> float:
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100.0 - (100.0 / (1.0 + rs))


def kd(
    highs: List[Optional[float]],
    lows: List[Optional[float]],
    closes: List[Optional[float]],
    n: int = 9,
) -> Tuple[List[Optional[float]], List[Optional[float]]]:
    """
    Standard KD (9,3,3) using intraday High/Low for RSV (consistent
    with most pro charting tools, e.g. TradingView):
        RSV = (Close - LowestLow_n) / (HighestHigh_n - LowestLow_n) * 100
        K   = 2/3 * prevK + 1/3 * RSV    (seed K = 50)
        D   = 2/3 * prevD + 1/3 * K      (seed D = 50)
    Output K/D start at index n-1 (the first day with a complete n-window).
    """
    length = len(closes)
    k_out: List[Optional[float]] = [None] * length
    d_out: List[Optional[float]] = [None] * length

    prev_k = 50.0
    prev_d = 50.0

    for i in range(length):
        if i < n - 1:
            continue
        h_window = highs[i - n + 1:i + 1]
        l_window = lows[i - n + 1:i + 1]
        if (closes[i] is None
                or any(v is None for v in h_window)
                or any(v is None for v in l_window)):
            continue
        hi = max(h_window)
        lo = min(l_window)
        if hi == lo:
            rsv = 50.0     # flat window → neutral
        else:
            rsv = (closes[i] - lo) / (hi - lo) * 100.0
        k = (2.0 / 3.0) * prev_k + (1.0 / 3.0) * rsv
        d = (2.0 / 3.0) * prev_d + (1.0 / 3.0) * k
        k_out[i] = k
        d_out[i] = d
        prev_k, prev_d = k, d
    return k_out, d_out


def ema(values: List[Optional[float]], n: int) -> List[Optional[float]]:
    """
    EMA seeded with the SMA of the first n values, then standard
    EMA recursion: ema = prev_ema * (1 - alpha) + value * alpha,
    where alpha = 2 / (n + 1).
    """
    length = len(values)
    out: List[Optional[float]] = [None] * length
    if length < n:
        return out

    # Find the first window of n consecutive non-None values.
    start = None
    for i in range(length - n + 1):
        if all(v is not None for v in values[i:i + n]):
            start = i + n - 1
            break
    if start is None:
        return out

    out[start] = sum(values[start - n + 1:start + 1]) / n  # SMA seed
    alpha = 2.0 / (n + 1)
    for i in range(start + 1, length):
        v = values[i]
        if v is None or out[i - 1] is None:
            out[i] = out[i - 1]
            continue
        out[i] = out[i - 1] * (1 - alpha) + v * alpha
    return out


def macd(closes: List[Optional[float]],
         fast: int = 12, slow: int = 26, signal: int = 9
         ) -> Tuple[List[Optional[float]], List[Optional[float]], List[Optional[float]]]:
    """Returns (DIF, signal-line, OSC) where OSC = DIF - signal."""
    ema_fast = ema(closes, fast)
    ema_slow = ema(closes, slow)
    dif = [
        (a - b) if (a is not None and b is not None) else None
        for a, b in zip(ema_fast, ema_slow)
    ]
    sig = ema(dif, signal)
    osc = [
        (a - b) if (a is not None and b is not None) else None
        for a, b in zip(dif, sig)
    ]
    return dif, sig, osc


def pct_change(values: List[Optional[float]]) -> List[Optional[float]]:
    """Day-over-day percentage change in percent (e.g. 1.23 means +1.23%)."""
    out: List[Optional[float]] = [None] * len(values)
    for i in range(1, len(values)):
        prev, curr = values[i - 1], values[i]
        if prev is None or curr is None or prev == 0:
            continue
        out[i] = (curr - prev) / prev * 100.0
    return out


# ---- Time-series assembly --------------------------------------------------

def parse_stock_rows(rows: List[List[str]]) -> List[Dict]:
    """Convert raw STOCK_DAY rows into list of dicts (oldest -> newest)."""
    out = []
    for r in rows:
        d = roc_to_western_date(r[STOCK_IDX_DATE])
        if d is None:
            continue
        shares = to_number(r[STOCK_IDX_VOLUME], integer=True)
        out.append({
            "date": d,
            "high": to_number(r[STOCK_IDX_HIGH]),
            "low": to_number(r[STOCK_IDX_LOW]),
            "close": to_number(r[STOCK_IDX_CLOSE]),
            "lots": (shares // 1000) if shares is not None else None,
        })
    return out


def parse_taiex_rows(rows: List[List[str]]) -> Dict[date, float]:
    """Map of date -> TAIEX close."""
    out: Dict[date, float] = {}
    for r in rows:
        d = roc_to_western_date(r[TAIEX_IDX_DATE])
        c = to_number(r[TAIEX_IDX_CLOSE])
        if d is not None and c is not None:
            out[d] = c
    return out


# ---- Excel output ----------------------------------------------------------

HEADER = [
    "Date", "TAIEX", "High", "Low", "Close", "Volume (lots)", "Change %",
    "MA5", "MA10", "MA20", "RSI(6)", "RSI(12)",
    "KD-K", "KD-D", "MACD-OSC",
]

# Column-by-column number formats (1-indexed to match openpyxl).
COL_FORMATS = {
    1:  "yyyy-mm-dd",            # Date
    2:  "#,##0.00",              # TAIEX
    3:  "#,##0.00",              # High
    4:  "#,##0.00",              # Low
    5:  "#,##0.00",              # Close
    6:  "#,##0",                 # Volume (lots)
    7:  "0.00;[Red]-0.00",       # Change %
    8:  "0.00",                  # MA5
    9:  "0.00",                  # MA10
    10: "0.00",                  # MA20
    11: "0.00",                  # RSI(6)
    12: "0.00",                  # RSI(12)
    13: "0.00",                  # KD-K
    14: "0.00",                  # KD-D
    15: "0.00;[Red]-0.00",       # MACD-OSC
}
COL_WIDTHS = [12, 11, 9, 9, 9, 13, 10, 9, 9, 9, 10, 10, 9, 9, 11]


def _add_or_replace_sheet(wb: Workbook, title: str):
    """
    Return a worksheet named `title`, ready to be filled.

    Behavior:
      * If a sheet with this name already exists, its contents are
        cleared in place — the sheet object itself is kept, so its
        position is preserved and any external references (formulas
        on other sheets like `=2324!B5`) keep working.
      * If not, a new sheet is appended at the end of the workbook.
      * The default empty 'Sheet' that openpyxl creates for a brand-new
        workbook is reused (renamed) for the very first stock, so we
        don't end up with a stray empty sheet.
    """
    # Reuse the auto-created blank "Sheet" of a fresh Workbook for the
    # first call only (single sheet, untouched, default name).
    if (
        len(wb.sheetnames) == 1
        and wb.sheetnames[0] == "Sheet"
        and wb.active.max_row == 1
        and wb.active.max_column == 1
    ):
        ws = wb.active
        ws.title = title
        return ws

    # Existing sheet — clear all rows and reset some sheet-level state
    # while keeping the sheet object itself (so its position and any
    # external formula references survive).
    if title in wb.sheetnames:
        ws = wb[title]
        # Drop every row. ws.max_row is 1 for a freshly-created sheet,
        # so guard against the no-op delete that would otherwise raise.
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)
        # Wipe per-sheet UI state that we re-apply below for fresh sheets.
        ws.auto_filter.ref = None
        ws.freeze_panes = None
        # Wipe column-level formatting (widths) so old layouts don't
        # bleed through if the column count ever changes.
        ws.column_dimensions.clear()
        return ws

    # Brand-new sheet at the end.
    return wb.create_sheet(title=title)


def write_stock_sheet(
    wb: Workbook,
    stock_no: str,
    series: List[Dict],
    taiex_close: Dict[date, float],
    output_rows: int,
) -> None:
    ws = _add_or_replace_sheet(wb, stock_no)

    # ---- Compute indicators on the FULL series (incl. warmup) ------------
    closes = [pt["close"] for pt in series]
    highs  = [pt["high"]  for pt in series]
    lows   = [pt["low"]   for pt in series]

    ma5  = sma(closes, 5)
    ma10 = sma(closes, 10)
    ma20 = sma(closes, 20)
    rsi6  = rsi_wilder(closes, 6)
    rsi12 = rsi_wilder(closes, 12)
    k_vals, d_vals = kd(highs, lows, closes, 9)
    _dif, _sig, osc = macd(closes, 12, 26, 9)
    chg_pct = pct_change(closes)

    # ---- Trim to the last `output_rows` rows for output -----------------
    start = max(0, len(series) - output_rows)

    # Header
    ws.append(HEADER)
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    for col_idx in range(1, len(HEADER) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold
        c.alignment = center

    # Data rows
    for i in range(start, len(series)):
        pt = series[i]
        ws.append([
            pt["date"],
            taiex_close.get(pt["date"]),
            pt["high"],
            pt["low"],
            pt["close"],
            pt["lots"],
            chg_pct[i],
            ma5[i],
            ma10[i],
            ma20[i],
            rsi6[i],
            rsi12[i],
            k_vals[i],
            d_vals[i],
            osc[i],
        ])

    # Formatting
    last_row = ws.max_row
    if last_row >= 2:
        for r in range(2, last_row + 1):
            for col, fmt in COL_FORMATS.items():
                ws.cell(row=r, column=col).number_format = fmt

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---- CLI -------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Fetch TWSE daily data + technical indicators for stocks."
    )
    p.add_argument(
        "--stocks", "-s",
        nargs="+",
        default=["2324"],
        help="One or more TWSE stock codes. Default: 2324",
    )
    p.add_argument(
        "--rows", "-n",
        type=int,
        default=30,
        help="Number of most recent trading days to OUTPUT per stock. Default: 30",
    )
    p.add_argument(
        "--output", "-o",
        default=None,
        help="Output .xlsx filename. Default: twse_<stocks>_<rows>d_<YYYYMMDD>.xlsx",
    )
    p.add_argument(
        "--no-taiex",
        action="store_true",
        help="Skip TAIEX column (faster).",
    )
    return p.parse_args()


def default_output_name(stocks: List[str], rows: int) -> str:
    today = datetime.now().strftime("%Y%m%d")
    joined = "_".join(stocks)
    return f"twse_{joined}_{rows}d_{today}.xlsx"


def main() -> None:
    args = parse_args()
    if args.rows <= 0:
        print("--rows must be a positive integer", file=sys.stderr)
        sys.exit(1)

    output_file = args.output or default_output_name(args.stocks, args.rows)
    fetch_count = args.rows + WARMUP_DAYS

    # Open existing file (merge-mode) if it's there, otherwise start fresh.
    if os.path.exists(output_file):
        try:
            wb = load_workbook(output_file)
            existing = ", ".join(wb.sheetnames) or "(none)"
            print(f"Opening existing file: {output_file}")
            print(f"  Existing sheets: {existing}")
            print(f"  Sheets matching --stocks will be replaced; others kept as-is.")
        except Exception as exc:
            print(f"Could not open existing {output_file} ({exc}); creating new one.",
                  file=sys.stderr)
            wb = Workbook()
    else:
        wb = Workbook()

    print(f"Output: last {args.rows} trading days "
          f"(fetching {fetch_count} including {WARMUP_DAYS}-day warmup)")

    # Fetch TAIEX once (shared across all stock sheets)
    taiex_close: Dict[date, float] = {}
    if not args.no_taiex:
        print("\n[TAIEX]")
        taiex_rows = fetch_recent_rows(fetch_taiex_month, fetch_count, "TAIEX")
        taiex_close = parse_taiex_rows(taiex_rows)

    # Per-stock sheets
    for stock_no in args.stocks:
        print(f"\n[{stock_no}]")
        rows = fetch_recent_rows(
            lambda y, m, s=stock_no: fetch_stock_month(y, m, s),
            fetch_count,
            stock_no,
        )
        if not rows:
            print(f"    [warn] no data collected for {stock_no}")
            _add_or_replace_sheet(wb, stock_no).append(HEADER)
            continue
        series = parse_stock_rows(rows)
        write_stock_sheet(wb, stock_no, series, taiex_close, args.rows)

    wb.save(output_file)
    print(f"\nDone. Saved to {output_file}")


if __name__ == "__main__":
    main()
